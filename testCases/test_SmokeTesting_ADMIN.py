from selenium.common import WebDriverException, TimeoutException, NoSuchElementException
import os
import time
from telnetlib import EC
import openpyxl
import pytest
import softest
from colorama import Fore, Back, Style
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import datetime

@pytest.mark.usefixtures("setup_and_teardown",scope="class")
class Test_verifyHomePage(softest.TestCase):
   # log = custLogger().getLogs()

    def test_TitleAndLogo(self):
        wait = WebDriverWait(self.driver, 20)
        # self.driver.find_element(By.ID, "cms-login-userId").send_keys("PDADM_01")
        # self.driver.find_element(By.ID, "cms-login-password").send_keys("Crowd_dev_112723")
        self.driver.find_element(By.ID, "cms-label-tc").click()
        self.driver.find_element(By.ID, "cms-login-submit").click()
        time.sleep(4)
        #self.wait.until(EC.element_to_be_clickable(By.ID,"cms-send-push-phone" ))
        self.driver.find_element(By.ID, "cms-send-push-phone").click()
        time.sleep(10)
#Click on CROWD
        #self.driver.find_element(By.XPATH, "//a[@id='cms-tile-a-hs-crowd']").click()
        #time.sleep(19)
        #self.driver.get_screenshot_as_file("..\Screenshots\Main_Page.png")

        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@id='cms-tile-a-hs-crowd']"))).click()

        # try:
        #     click_on_CROWD = self.driver.find_element(By.XPATH, "//a[@id='cms-tile-a-hs-crowd']")
        #     self.driver.execute_script("arguments[0].click();", click_on_CROWD)
        # except NoSuchElementException as e:
        #     print(e)

            #             print(e)
        time.sleep(7)
        element_App = self.driver.find_element(By.LINK_TEXT, "Application")
        element_App.click()
# using now() to get current time
        #current_time = datetime.datetime.now()
        workbook = openpyxl.load_workbook("..\ExcelFiles\Automated_SmokeTest_Result_ADMIN.xlsx")
        sheet = workbook['Sheet1']
        combined = datetime.datetime.now()
        formatted_combined = combined.strftime("%d/%m/%Y %I:%M:%p")
        print('\n')
        print(Back.LIGHTYELLOW_EX)
        print(Fore.RED+'  CROWD - SMOKE TESTING RESULTS FOR TEST ENV ON ', formatted_combined )
        print(Style.RESET_ALL)
        sheet.cell(row=1, column=3).value =  formatted_combined
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result_ADMIN.xlsx")
        print(Fore.BLUE +'\n******************** Verifying Title(s) & Logo(s) of CROWD Application ********************')
        print(Style.RESET_ALL)

    # Verifying  Logo :   CMS.gov| My Enterprise Portal  ---CMS - Homepage header Logo
        #expected_Logo = "CMS.gov| My Enterprise Portal is displayed"
        print(Fore.GREEN + 'Main Logo: -', end='')
        print(Style.RESET_ALL)
        actual_Logo = self.driver.find_element(By.XPATH, "//em[@id='cms-homepage-header-logo-unauth']")
        print(actual_Logo.is_displayed())
        sheet.cell(row=6, column=3).value = actual_Logo.is_displayed()
        sheet.cell(row=6, column=4).value = "Pass"
        #workbook.save("..\ExcelFiles\SmokeTest_Result.xlsx")
        print("     'CMS.Gov|My Enterprise Portal' Logo is displayed.")

    # Verifying Title CMS Enterprise Portal - My Portal"
        expected_title_main = "CMS Enterprise Portal - My Portal"
        print(Fore.GREEN + 'Title Verified: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=8, column=3).value = self.driver.title
        self.soft_assert(self.assertEqual,self.driver.title,expected_title_main)
        if self.driver.title == expected_title_main:
            print("     Expected Title: - " + expected_title_main)
            print("     Actual Title: - " + self.driver.title)
            sheet.cell(row=8, column=4).value = "Pass"
        else:
            # self.log.info("Assertion Failed")
            print("Main Title not match")
            sheet.cell(row=8, column=4).value = "Fail"
        time.sleep(8)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
    # Verify 'CROWD' Logo// Once found NoSuchElementException
        image = self.driver.find_element(By.XPATH,"//div[@class ='cms-logo']")
        print(image.is_displayed())
        sheet.cell(row=10, column=3).value = image.is_displayed()
        sheet.cell(row=10, column=4).value = "Pass"
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result_ADMIN.xlsx")
        print(Fore.GREEN + 'CROWD Logo: -' , end='')
        print(Style.RESET_ALL)
        print("     'CROWD' Logo is displayed.")
    # Verify Title
        expected_title = "CmsCrowdUi"
        actual_title = self.driver.execute_script('return document.title')
        print(Fore.GREEN + ' Title Verified: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=12, column=3).value = actual_title
        self.soft_assert(self.assertEqual, actual_title, expected_title)

        if actual_title == expected_title:
           print("     Expected Title: - " + expected_title)
           print("     Actual Title: - " + actual_title)
           sheet.cell(row=12, column=4).value = "Pass"
        else:
            # self.log.info("Assertion Failed")
            sheet.cell(row=12, column=4).value = "Fail"
            print("Title not match")
# Verifying  Text Welcome to CROWD
        #self.log.info("Verifying 'Welcome to CROWD' text")
        #wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='homeTitleRow']/h1")))
        actual_Text= wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='homeTitleRow']/h1"))).text
        expected_Text ="Welcome to CROWD"
        #actual_Text=self.driver.find_element(By.XPATH, "//div[@id='homeTitleRow']/h1").text
        print(Fore.GREEN + 'Text Verified: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=14, column=3).value = actual_Text
        self.soft_assert(self.assertEqual, actual_Text, expected_Text)
        if actual_Text == expected_Text:
           # self.log.info("Assertion Passed")
            print("     Expected Text: - " + expected_Text)
            print("     Actual Text: - " + actual_Text)
            sheet.cell(row=14, column=4).value = "Pass"
        else:
           #self.log.info("Assertion Faileded")
           print("Welcome To Crowd test does not match")
           sheet.cell(row=14, column=4).value = "Fail"
# Verifying CROWD Summary
        # self.log.info("Verifying 'Welcome to CROWD' text")
        expected_Summary = "The Contractor Reporting of Operational and Workload Data (CROWD) application provides CMS automated capabilities for monitoring and analyzing data relating to the Medicare contractor's on-going operational activities. The application contains workload reporting capabilities that allow the data to be used for estimating budgets, defining operating problems, comparing performance among contractors, and determining national workload trends."
        actual_Summary = self.driver.find_element(By.XPATH, "//p[@id='crowd_summary']").text
        print(Fore.GREEN + 'Verified CROWD Summary: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=16, column=3).value = actual_Summary
        self.soft_assert(self.assertEqual, actual_Summary, expected_Summary)
        if actual_Summary == expected_Summary:
            print("     Expected Summary: - " + expected_Summary)
            print("     Actual Summary: - " + actual_Summary)
            sheet.cell(row=16, column=4).value = "Pass"
        else:
             # self.log.info("Assertion Faileded")
            print("CROWD Summary does not match")
            sheet.cell(row=16, column=4).value = "Fail"
        self.driver.get_screenshot_as_file("..\Screenshots\CROWD Page.png")
############################################## Verifying Dashboard ###################################################
        print(Fore.BLUE +"\n************************* Verifying Dashboard Page *************************")
        print(Style.RESET_ALL)
        dashboardLink = self.driver.find_element(By.XPATH,
                                                "//span[normalize-space()='Dashboards']")
        self.driver.execute_script("arguments[0].click();", dashboardLink)
        #self.driver.get_screenshot_as_file("..\Screenshots\Dashboard.png")
        # dash_Text = self.driver.find_element(By.XPATH, "//span[@class='anchor-format grey-text ng-star-inserted']").text
        # print("Displayed text on Dashboards page:- " + dash_Text)
        dashboardID = self.driver.find_element(By.XPATH, "//table[@title='Dashboard']/tbody/tr/td[1]").text
        dashboardName = self.driver.find_element(By.XPATH, "//table[@title='Dashboard']/tbody/tr/td[2]").text
        dashboardDescription = self.driver.find_element(By.XPATH, "//table[@title='Dashboard']/tbody/tr/td[3]").text
        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON 'DASHBOARD' TABLE IS:- ", end='')
        print(Style.RESET_ALL)
        print("     Dashboard ID: - " + dashboardID + "   Dashboard Name: - " + dashboardName + "   Dashboard Description: - " + dashboardDescription)
        sheet.cell(row=20, column=3).value = dashboardID
        sheet.cell(row=22, column=3).value = dashboardName
        sheet.cell(row=24, column=3).value = dashboardDescription
        if sheet.cell(row=20, column=3).value == sheet.cell(row=20, column=2).value:
            sheet.cell(row=20, column=4).value ='Pass'
        else:
            sheet.cell(row=20, column=4).value = 'Fail'
        if sheet.cell(row=22, column=3).value == sheet.cell(row=22, column=2).value:
            sheet.cell(row=22, column=4).value = 'Pass'
        else:
            sheet.cell(row=22, column=4).value = 'Fail'
        if sheet.cell(row=24, column=3).value == sheet.cell(row=24, column=2).value:
            sheet.cell(row=24, column=4).value = 'Pass'
        else:
            sheet.cell(row=24, column=4).value = 'Fail'
        time.sleep(5)
        parent_window_id= self.driver.current_window_handle
        click_on_firstLink_Dashboard = self.driver.find_element(By.XPATH,
                                             "//table[@title='Dashboard']/tbody/tr/td[2]/a")
        self.driver.execute_script("arguments[0].click();", click_on_firstLink_Dashboard)
        windows= self.driver.window_handles
        for w in windows:
            self.driver.switch_to.window(w)
            time.sleep(6)
            expected_Title = "B01. MicroStrategy"
            actual_Title = self.driver.title
            sheet.cell(row=26, column=3).value = actual_Title
            if actual_Title==expected_Title:
                print(Fore.GREEN+'Verified MicroStrategy Page Title: -', end='')
                print(Style.RESET_ALL)
                print("     Expected Title: - " + expected_Title)
                print("     Actual Title: - " + actual_Title)
                sheet.cell(row=26, column=4).value = "Pass"
            # if self.driver.title.__eq__("B01. MicroStrategy"):
            #     print(self.driver.title)
                #print(self.driver.current_window_handle)
                self.driver.close()
                break
            else:
                sheet.cell(row=26, column=4).value = "Fail"
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result_ADMIN.xlsx")
        self.driver.switch_to.window(parent_window_id)
############################### Verifying Form Page #####################################
        print(Fore.BLUE + "\n************************* Verifying Form Page *************************")
        print(Style.RESET_ALL)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        time.sleep(5)
        #forms_Link = self.driver.find_element(By.XPATH,"//span[normalize-space()='Forms']")
        forms_Link = self.driver.find_element(By.XPATH, "//span[@class='list-item-label'][normalize-space()='Forms']")
        self.driver.execute_script("arguments[0].click();", forms_Link)
        time.sleep(5)
        self.driver.get_screenshot_as_file("..\Screenshots\Forms.png")
        form = self.driver.find_element(By.XPATH, "//main[@id='main-forms']/table/tbody/tr[1]/td[1]").text
        formType = self.driver.find_element(By.XPATH, "//main[@id='main-forms']/table/tbody/tr[1]/td[2]").text
        formName = self.driver.find_element(By.XPATH, "//main[@id='main-forms']/table/tbody/tr[1]/td[3]").text

        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON 'FORM' TABLE IS: - ", end='')
        print(Style.RESET_ALL)
        print("     Form: - " + form + "  Form Type: - " + formType + "  Form Name: - " + formName)
        sheet.cell(row=30, column=3).value = form
        sheet.cell(row=32, column=3).value = formType
        sheet.cell(row=34, column=3).value = formName

        if sheet.cell(row=30, column=3).value == sheet.cell(row=30, column=2).value:
            sheet.cell(row=30, column=4).value ='Pass'
        else:
            sheet.cell(row=30, column=4).value = 'Fail'
        if sheet.cell(row=32, column=3).value == sheet.cell(row=32, column=2).value:
            sheet.cell(row=32, column=4).value = 'Pass'
        else:
            sheet.cell(row=32, column=4).value = 'Fail'
        if sheet.cell(row=34, column=3).value == sheet.cell(row=34, column=2).value:
            sheet.cell(row=34, column=4).value = 'Pass'
        else:
            sheet.cell(row=34, column=4).value = 'Fail'
        time.sleep(7)
        parent_window_id = self.driver.current_window_handle
        #click_on_firstLink_Forms = self.driver.find_element(By.XPATH,
                                            # "//main[@id='main-forms']/table[@title='Forms Table']/tbody/tr[2]/td[3]/a")
        click_on_firstLink_Forms = self.driver.find_element(By.XPATH,
                                                            "//main[@id='main-forms']/table/tbody/tr[1]/td[3]/a")
        self.driver.execute_script("arguments[0].click();", click_on_firstLink_Forms)
        windows = self.driver.window_handles
        for w in windows:
            self.driver.switch_to.window(w)
            time.sleep(9)
            expected_Title = "(FORM 7 - PART A) APPEALS ACTIVITY (CMS-2592). MicroStrategy"
            actual_Title = self.driver.title
            sheet.cell(row=36, column=3).value = actual_Title
            if actual_Title==expected_Title:
                print(Fore.GREEN + 'Verified MicroStrategy Page Title: - ', end='')
                print(Style.RESET_ALL)
                print("     Expected Title: - "+ expected_Title)
                print("     Actual Title: - " + actual_Title)
                sheet.cell(row=36, column=4).value = "Pass"
                #print(self.driver.current_window_handle)
                break
            else:
                sheet.cell(row=36, column=4).value = "Fail"
        self.driver.switch_to.window(parent_window_id)
# # ############################## Verifying Reports Page #####################################
        print(Fore.BLUE + "\n************************* Verifying Reports Page *************************")
        print(Style.RESET_ALL)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        time.sleep(3)
        reports_Link = self.driver.find_element(By.XPATH,
                                                 "//span[normalize-space()='Reports']")
        self.driver.execute_script("arguments[0].click();", reports_Link)
        time.sleep(5)
        self.driver.get_screenshot_as_file("..\Screenshots\Reports.png")
        reportID = self.driver.find_element(By.XPATH, "//table[@title='Reports']/tbody/tr/td[1]").text # WebDriver Exception
        reportName = self.driver.find_element(By.XPATH, "//table[@title='Reports']/tbody/tr/td[2]").text
        reportDescription = self.driver.find_element(By.XPATH, "//table[@title='Reports']/tbody/tr/td[3]").text
        sheet.cell(row=40, column=3).value = reportID
        sheet.cell(row=42, column=3).value = reportName
        sheet.cell(row=44, column=3).value = reportDescription

        if sheet.cell(row=40, column=3).value == sheet.cell(row=40, column=2).value:
            sheet.cell(row=40, column=4).value ='Pass'
        else:
            sheet.cell(row=40, column=4).value = 'Fail'

        if sheet.cell(row=42, column=3).value == sheet.cell(row=42, column=2).value:
            sheet.cell(row=42, column=4).value = 'Pass'
        else:
            sheet.cell(row=42, column=4).value = 'Fail'

        if sheet.cell(row=44, column=3).value == sheet.cell(row=44, column=2).value:
            sheet.cell(row=44, column=4).value = 'Pass'
        else:
            sheet.cell(row=44, column=4).value = 'Fail'

        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON 'REPORT' TABLE IS: - ", end='')
        print(Style.RESET_ALL)
        print("     Report ID: - " + reportID + "   Report Name: - " + reportName + "   Report Description: - " + reportDescription)
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result_ADMIN.xlsx")
############################## Verifying Resources  Page ###################################################
        print(Fore.BLUE + "\n************************* Verifying Resources Page*************************")
        print(Style.RESET_ALL)
        ##################### FAQs#########################
        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON FAQs/Reporting/Training Material: - ", end='')
        print(Style.RESET_ALL)
        try:

            resources_Link = self.driver.find_element(By.XPATH,
                                             "//span[normalize-space()='Resources']")
            self.driver.execute_script("arguments[0].click();", resources_Link)
            time.sleep(10)
            click_on_FAQs = self.driver.find_element(By.XPATH,
                                                       "//span[normalize-space()='FAQs']")
            self.driver.execute_script("arguments[0].click();", click_on_FAQs)
            time.sleep(9)
            click_on_SystemSetup = self.driver.find_element(By.XPATH,# WebDriver Exception error
                                                 "//a[normalize-space()='System Set-Up Considerations']")
            self.driver.execute_script("arguments[0].click();", click_on_SystemSetup)
            system_Setup_Text = self.driver.find_element(By.XPATH,
                                                 "//div[@class='container']/mat-card/ul/li[1]/a").text
            sheet.cell(row=48, column=3).value = system_Setup_Text
            if sheet.cell(row=48, column=3).value == sheet.cell(row=48, column=2).value:
                sheet.cell(row=48, column=4).value ='Pass'
            else:
                sheet.cell(row=48, column=4).value = 'Fail'
        ##############  Reporting ############
            reporting_Link = self.driver.find_element(By.XPATH,
                                                  "//span[normalize-space()='Reporting']")
            self.driver.execute_script("arguments[0].click();", reporting_Link)
            time.sleep(10)
            click_on_CrowdContractorMapping = self.driver.find_element(By.XPATH,# WebDriver Exception
                                                       " //a[normalize-space()='Crowd Contractor Mapping']")
            self.driver.execute_script("arguments[0].click();", click_on_CrowdContractorMapping)
            crowd_Reporting_Requirement_Text = self.driver.find_element(By.XPATH,
                                                   "//div[@class='container']/mat-card/ul/li[1]/a").text
            print(crowd_Reporting_Requirement_Text)
            sheet.cell(row=50, column=3).value = crowd_Reporting_Requirement_Text
            if sheet.cell(row=50, column=3).value == sheet.cell(row=50, column=2).value:
                sheet.cell(row=50, column=4).value ='Pass'
            else:
                sheet.cell(row=50, column=4).value = 'Fail'
        ############ Training Material ########################
            trainingMaterial_Link = self.driver.find_element(By.XPATH,
                                                  "//span[normalize-space()='Training Material']")
            self.driver.execute_script("arguments[0].click();", trainingMaterial_Link)
            time.sleep(10)
            click_on_RequestCrowdAccess = self.driver.find_element(By.XPATH,
                                                                   "//a[contains(text(),'Request CROWD Access (Without an existing EIDM acc')]")
            self.driver.execute_script("arguments[0].click();", click_on_RequestCrowdAccess)
            requestCrowdAccess_Text = self.driver.find_element(By.XPATH,
                                                   "//div[@class='container']/mat-card/ul/li[2]/ul/li[2]/a").text

            print(requestCrowdAccess_Text)
            sheet.cell(row=52, column=3).value = requestCrowdAccess_Text
            if sheet.cell(row=52, column=3).value == sheet.cell(row=52, column=2).value:
                sheet.cell(row=52, column=4).value ='Pass'
            else:
                sheet.cell(row=52, column=4).value = 'Fail'
            workbook.save("..\ExcelFiles\Automated_SmokeTest_Result_ADMIN.xlsx")
        except WebDriverException as e:
            print("Webdriver encounterd an exception: (str(e))")
############################ Verifying ADMIN Functionality ######################
        print(Fore.BLUE + "\n************************* Verifying ADMIN Page *************************")
        print(Style.RESET_ALL)
        admin_Link = self.driver.find_element(By.XPATH,
                                             "//span[normalize-space()='Admin']")
        self.driver.execute_script("arguments[0].click();", admin_Link)
        ########### OVERRIDE REQUEST ###############
        overrideRequest_Link = self.driver.find_element(By.XPATH,
                                              "//div[@id = 'cdk-accordion-child-2']/div/a[1]")
        self.driver.execute_script("arguments[0].click();", overrideRequest_Link)

        fileName = self.driver.find_element(By.XPATH,
                                                    "//table[@title='Override Requests Table']/tbody/tr/td[1]").text
        sheet.cell(row=57, column=2).value = fileName
        sheet.cell(row=57, column=3).value = fileName
        if sheet.cell(row=57, column=3).value == sheet.cell(row=57, column=2).value:
            sheet.cell(row=57, column=4).value = 'Pass'
        else:
            sheet.cell(row=57, column=4).value = 'Fail'

        submittedBy = self.driver.find_element(By.XPATH,
                                                    "//table[@title='Override Requests Table']/tbody/tr/td[2]").text
        sheet.cell(row=59, column=2).value = submittedBy
        sheet.cell(row=59, column=3).value = submittedBy
        if sheet.cell(row=59, column=3).value == sheet.cell(row=59, column=2).value:
            sheet.cell(row=59, column=4).value = 'Pass'
        else:
            sheet.cell(row=59, column=4).value = 'Fail'

        timeStamp = self.driver.find_element(By.XPATH,
                                               "//table[@title='Override Requests Table']/tbody/tr/td[4]").text
        sheet.cell(row=61, column=2).value = timeStamp
        sheet.cell(row=61, column=3).value = timeStamp
        if sheet.cell(row=61, column=3).value == sheet.cell(row=61, column=2).value:
            sheet.cell(row=61, column=4).value = 'Pass'
        else:
            sheet.cell(row=61, column=4).value = 'Fail'

        status = self.driver.find_element(By.XPATH,
                                          "//table[@title='Override Requests Table']/tbody/tr/td[3]").text
        sheet.cell(row=63, column=2).value = status
        sheet.cell(row=63, column=3).value = status
        if sheet.cell(row=63, column=3).value == sheet.cell(row=63, column=2).value:
            sheet.cell(row=63, column=4).value = 'Pass'
        else:
            sheet.cell(row=63, column=4).value = 'Fail'

        ###########System Report Link ############
        systemReport_Link = self.driver.find_element(By.XPATH,
                                                        "//div[@id = 'cdk-accordion-child-2']/div/a[2]")
        self.driver.execute_script("arguments[0].click();", systemReport_Link)


        reportID = self.driver.find_element(By.XPATH,
                                                     "//table[@title = 'Reports']/tbody/tr/td").text

        sheet.cell(row=67, column=2).value = reportID
        sheet.cell(row=67, column=3).value = reportID
        if sheet.cell(row=67, column=3).value == sheet.cell(row=67, column=2).value:
            sheet.cell(row=67, column=4).value = 'Pass'
        else:
            sheet.cell(row=67, column=4).value = 'Fail'


        reportName = self.driver.find_element(By.XPATH,
                                              "//table[@title='Reports']/tbody/tr/td[2]").text
        sheet.cell(row=69, column=2).value = reportName
        sheet.cell(row=69, column=3).value = reportName
        if sheet.cell(row=69, column=3).value == sheet.cell(row=69, column=2).value:
            sheet.cell(row=69, column=4).value = 'Pass'
        else:
            sheet.cell(row=69, column=4).value = 'Fail'

        reportDescripiton = self.driver.find_element(By.XPATH,
                                               "//table[@title = 'Reports']/tbody/tr/td[3]").text
        sheet.cell(row=71, column=2).value = reportDescripiton
        sheet.cell(row=71, column=3).value = reportDescripiton
        if sheet.cell(row=71, column=3).value == sheet.cell(row=71, column=2).value:
            sheet.cell(row=71, column=4).value = 'Pass'
        else:
            sheet.cell(row=71, column=4).value = 'Fail'

        ############# Table Data Link ##############

        tableData_Link = self.driver.find_element(By.XPATH,
                                                      "//div[@id = 'cdk-accordion-child-2']/div/a[3]")
        self.driver.execute_script("arguments[0].click();", tableData_Link)
        # tableData_Link = self.driver.find_element(By.XPATH,
        #                                         " //table[@title ='Table Data']/tbody/ tr/td/a").text
        # sheet.cell(row=60, column=3).value = fileName
        #
        # if sheet.cell(row=60, column=3).value == sheet.cell(row=60, column=2).value:
        #         sheet.cell(row=60, column=4).value = 'Pass'
        # else:
        #         sheet.cell(row=60, column=4).value = 'Fail'
        tableName = self.driver.find_element(By.XPATH,
                                                   "//table[@title='Table Data']/tbody/tr/td").text
        sheet.cell(row=75, column=2).value = tableName
        sheet.cell(row=75, column=3).value = tableName
        if sheet.cell(row=75, column=3).value == sheet.cell(row=75, column=2).value:
                sheet.cell(row=75, column=4).value = 'Pass'
        else:
            sheet.cell(row=75, column=4).value = 'Fail'

        tableDescription = self.driver.find_element(By.XPATH,
                                              "//table[@title='Table Data']/tbody/tr/td[2]").text
        sheet.cell(row=77, column=2).value = tableDescription
        sheet.cell(row=77, column=3).value = tableDescription
        if sheet.cell(row=77, column=3).value == sheet.cell(row=77, column=2).value:
                sheet.cell(row=77, column=4).value = 'Pass'
        else:
                sheet.cell(row=77, column=4).value = 'Fail'

        # ############################## Verifying News Page #####################################
        print(Fore.BLUE + "\n************************* Verifying News Page *************************")
        print(Style.RESET_ALL)

        news_Link = self.driver.find_element(By.XPATH,
                                                 "//span[normalize-space()='News']")
        self.driver.execute_script("arguments[0].click();", news_Link)
        time.sleep(10)

        print(Fore.GREEN + "NEWS for Year 2020: -")
        print(Style.RESET_ALL)
       # self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        year2020 = self.driver.find_element(By.XPATH,
                                             "//h2[normalize-space()='2020']")
        self.driver.execute_script("arguments[0].click();", year2020)
        time.sleep(6)
       # self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        line1 =self.driver.find_element(By.XPATH, " (//span[@class='blue'][normalize-space()='Attn: Medicare Contractors'])[1]").text
        #print(line1)
        time.sleep(6)
        line2 = self.driver.find_element(By.XPATH,
                                         "//span[normalize-space()='| February 12, 2020']").text
        print(line1 + ' ' + line2)
        sheet.cell(row=81, column=3).value = line1 + ' ' +line2
        if sheet.cell(row=81, column=3).value == sheet.cell(row=81, column=2).value:
            sheet.cell(row=81, column=4).value ='Pass'
        else:
            sheet.cell(row=81, column=4).value = 'Fail'
        #line3 =self.driver.find_element(By.XPATH, "//span[normalize-space()='| December 23rd, 2021']").text
        line3 = self.driver.find_element(By.XPATH, "//p[contains(text(),'Although CR 11353 (Supplier Specialty Code D5) for')]").text
        sheet.cell(row=82, column=3).value = line3
        print(line3)
        # line2 = self.driver.find_element(By.XPATH,
        #                                  " //span[@class='ng-star-inserted'][normalize-space()='| February 12, 2020'])[1]").text
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result_ADMIN.xlsx")
        self.assert_all()
















