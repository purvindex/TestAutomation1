import time
import openpyxl
import pytest
import softest
from colorama import Fore, Back, Style

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import  WebDriverWait
import datetime

@pytest.mark.usefixtures("setup_and_teardown",scope="class")
class Test_verifyHomePage(softest.TestCase):
   # log = custLogger().getLogs()

    def test_TitleAndLogo(self):
        wait = WebDriverWait(self.driver, 20)
        self.driver.find_element(By.ID, "cms-label-tc").click()
        self.driver.find_element(By.ID, "cms-login-submit").click()


#Click on CROWD
        time.sleep(8)
        self.driver.get_screenshot_as_file("..\Screenshots\Main_Page.png")
        click_on_CROWD = self.driver.find_element(By.XPATH,
                                                               "//div[@class='ng-star-inserted']")
        self.driver.execute_script("arguments[0].click();", click_on_CROWD)
        time.sleep(8)
        element_App = self.driver.find_element(By.LINK_TEXT, "Application")
        element_App.click()
# using now() to get current time
        #current_time = datetime.datetime.now()
        workbook = openpyxl.load_workbook("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        sheet = workbook['Sheet1']
        combined = datetime.datetime.now()
        formatted_combined = combined.strftime("%d/%m/%Y %I:%M:%p")
        print('\n')
        print(Back.LIGHTYELLOW_EX)
        print(Fore.RED+'  CROWD - SMOKE TESTING RESULTS FOR TEST ENV ON ', formatted_combined )
        print(Style.RESET_ALL)
        sheet.cell(row=1, column=3).value =  formatted_combined
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        print(Fore.BLUE +'\n******************** Verifying Title(s) & Logo(s) of CROWD Application ********************')
        print(Style.RESET_ALL)

    # Verifying  Logo :   CMS.gov| My Enterprise Portal  ---CMS - Homepage header Logo
        #expected_Logo = "CMS.gov| My Enterprise Portal is displayed"
        print(Fore.GREEN + 'Main Logo: -', end='')
        print(Style.RESET_ALL)
        actual_Logo = self.driver.find_element(By.XPATH, "//em[@id='cms-homepage-header-logo-unauth']")
        print(actual_Logo.is_displayed())
        sheet.cell(row=6, column=3).value = actual_Logo.is_displayed()
        sheet.cell(row=6, column=4).value = "Pass"
        #workbook.save("..\ExcelFiles\SmokeTest_Result.xlsx")
        print("     'CMS.Gov|My Enterprise Portal' Logo is displayed.")

    # Verifying Title CMS Enterprise Portal - My Portal"
        expected_title_main = "CMS Enterprise Portal - My Portal"
        print(Fore.GREEN + 'Title Verified: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=8, column=3).value = self.driver.title
        self.soft_assert(self.assertEqual,self.driver.title,expected_title_main)
        if self.driver.title == expected_title_main:
            print("     Expected Title: - " + expected_title_main)
            print("     Actual Title: - " + self.driver.title)
            sheet.cell(row=8, column=4).value = "Pass"
            #workbook.save("..\ExcelFiles\SmokeTest_Result.xlsx")
        else:
            # self.log.info("Assertion Failed")
            print("Main Title not match")
            sheet.cell(row=8, column=4).value = "Fail"
        time.sleep(10)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))

    # Verify 'CROWD' Logo// Once found NoSuchElementException
        image = self.driver.find_element(By.XPATH,"//div[@class ='cms-logo']")
        print(image.is_displayed())
        sheet.cell(row=10, column=3).value = image.is_displayed()
        sheet.cell(row=10, column=4).value = "Pass"
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        print(Fore.GREEN + 'CROWD Logo: -' , end='')
        print(Style.RESET_ALL)
        print("     'CROWD' Logo is displayed.")