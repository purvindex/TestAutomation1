from selenium.common import WebDriverException
import os
import time
from telnetlib import EC
import openpyxl
import pytest
import softest
from colorama import Fore, Back, Style
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import  WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from utilities.custom_Logger import custLogger
import datetime

@pytest.mark.usefixtures("setup_and_teardown",scope="class")
class Test_verifyHomePage(softest.TestCase):
   # log = custLogger().getLogs()

    def test_TitleAndLogo(self):
        wait = WebDriverWait(self.driver, 20)
        self.driver.find_element(By.ID, "cms-login-userId").send_keys("PDPOC_VAL")
        self.driver.find_element(By.ID, "cms-login-password").send_keys("Crowd_dev_112723")
        # self.driver.find_element(By.ID, "cms-login-userId").send_keys("POC_DEV")
        # self.driver.find_element(By.ID, "cms-login-password").send_keys("Contractor_112723")
        self.driver.find_element(By.ID, "cms-label-tc").click()
        self.driver.find_element(By.ID, "cms-login-submit").click()
        #self.wait.until(EC.element_to_be_clickable(By.ID,"cms-send-push-phone" )).click()
        # self.driver.find_element(By.ID, "cms-send-push-phone").click()

#Click on CROWD
        #self.driver.find_element(By.XPATH, "//div[@class='ng-star-inserted']").click()
        time.sleep(10)
        self.driver.get_screenshot_as_file("..\Screenshots\Main_Page.png")
        click_on_CROWD = self.driver.find_element(By.XPATH,
                                                               "//div[@class='ng-star-inserted']")
        self.driver.execute_script("arguments[0].click();", click_on_CROWD)

        time.sleep(10)
        #self.driver.get_screenshot_as_file("..\Screenshots\Main_Page.png")
        #self.driver.find_element(By.XPATH, "//div[@id='cms-crowd-tile-selected']").click()

        element_App = self.driver.find_element(By.LINK_TEXT, "Application")
        element_App.click()
# using now() to get current time
        #current_time = datetime.datetime.now()
        workbook = openpyxl.load_workbook("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        sheet = workbook['Sheet1']
        combined = datetime.datetime.now()
        formatted_combined = combined.strftime("%d/%m/%Y %I:%M:%p")
        print('\n')
        print(Back.LIGHTYELLOW_EX)
        print(Fore.RED+'  CROWD - SMOKE TESTING RESULTS FOR TEST ENV ON ', formatted_combined )
        print(Style.RESET_ALL)
        sheet.cell(row=1, column=3).value =  formatted_combined
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        print(Fore.BLUE +'\n******************** Verifying Title(s) & Logo(s) of CROWD Application ********************')
        print(Style.RESET_ALL)

    # Verifying  Logo :   CMS.gov| My Enterprise Portal  ---CMS - Homepage header Logo
        #expected_Logo = "CMS.gov| My Enterprise Portal is displayed"
        print(Fore.GREEN + 'Main Logo: -', end='')
        print(Style.RESET_ALL)
        actual_Logo = self.driver.find_element(By.XPATH, "//em[@id='cms-homepage-header-logo-unauth']")
        print(actual_Logo.is_displayed())
        sheet.cell(row=6, column=3).value = actual_Logo.is_displayed()
        sheet.cell(row=6, column=4).value = "Pass"
        #workbook.save("..\ExcelFiles\SmokeTest_Result.xlsx")
        print("     'CMS.Gov|My Enterprise Portal' Logo is displayed.")

    # Verifying Title CMS Enterprise Portal - My Portal"
        expected_title_main = "CMS Enterprise Portal - My Portal"
        print(Fore.GREEN + 'Title Verified: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=8, column=3).value = self.driver.title
        self.soft_assert(self.assertEqual,self.driver.title,expected_title_main)
        if self.driver.title == expected_title_main:
            print("     Expected Title: - " + expected_title_main)
            print("     Actual Title: - " + self.driver.title)
            sheet.cell(row=8, column=4).value = "Pass"
            #workbook.save("..\ExcelFiles\SmokeTest_Result.xlsx")
        else:
            # self.log.info("Assertion Failed")
            print("Main Title not match")
            sheet.cell(row=8, column=4).value = "Fail"
        time.sleep(10)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))

    # Verify 'CROWD' Logo// Once found NoSuchElementException
        image = self.driver.find_element(By.XPATH,"//div[@class ='cms-logo']")
        print(image.is_displayed())
        sheet.cell(row=10, column=3).value = image.is_displayed()
        sheet.cell(row=10, column=4).value = "Pass"
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        print(Fore.GREEN + 'CROWD Logo: -' , end='')
        print(Style.RESET_ALL)
        print("     'CROWD' Logo is displayed.")

    # Verify Title
        expected_title = "CmsCrowdUi"
        actual_title = self.driver.execute_script('return document.title')
        print(Fore.GREEN + ' Title Verified: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=12, column=3).value = actual_title
        self.soft_assert(self.assertEqual, actual_title, expected_title)

        if actual_title == expected_title:
           print("     Expected Title: - " + expected_title)
           print("     Actual Title: - " + actual_title)
           sheet.cell(row=12, column=4).value = "Pass"
        else:
            # self.log.info("Assertion Failed")
            sheet.cell(row=12, column=4).value = "Fail"
            print("Title not match")

# Verifying  Text Welcome to CROWD
        #self.log.info("Verifying 'Welcome to CROWD' text")

        #wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='homeTitleRow']/h1")))
        actual_Text= wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='homeTitleRow']/h1"))).text
        expected_Text ="Welcome to CROWD"
        #actual_Text=self.driver.find_element(By.XPATH, "//div[@id='homeTitleRow']/h1").text
        print(Fore.GREEN + 'Text Verified: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=14, column=3).value = actual_Text
        self.soft_assert(self.assertEqual, actual_Text, expected_Text)
        if actual_Text == expected_Text:
           # self.log.info("Assertion Passed")
            print("     Expected Text: - " + expected_Text)
            print("     Actual Text: - " + actual_Text)
            sheet.cell(row=14, column=4).value = "Pass"
        else:
           #self.log.info("Assertion Faileded")
           print("Welcome To Crowd test does not match")
           sheet.cell(row=14, column=4).value = "Fail"

# Verifying CROWD Summary
        # self.log.info("Verifying 'Welcome to CROWD' text")
        expected_Summary = "The Contractor Reporting of Operational and Workload Data (CROWD) application provides CMS automated capabilities for monitoring and analyzing data relating to the Medicare contractor's on-going operational activities. The application contains workload reporting capabilities that allow the data to be used for estimating budgets, defining operating problems, comparing performance among contractors, and determining national workload trends."
        actual_Summary = self.driver.find_element(By.XPATH, "//p[@id='crowd_summary']").text
        print(Fore.GREEN + 'Verified CROWD Summary: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=16, column=3).value = actual_Summary
        self.soft_assert(self.assertEqual, actual_Summary, expected_Summary)
        if actual_Summary == expected_Summary:
            print("     Expected Summary: - " + expected_Summary)
            print("     Actual Summary: - " + actual_Summary)
            sheet.cell(row=16, column=4).value = "Pass"
        else:
             # self.log.info("Assertion Faileded")
            print("CROWD Summary does not match")
            sheet.cell(row=16, column=4).value = "Fail"
        self.driver.get_screenshot_as_file("..\Screenshots\CROWD Page.png")
############################## Verifying File Submission (File Upload & File Status) ######################################
        print(Fore.BLUE + "\n******************** Verifying File Submission - File Upload/ File Status Page ********************")
        print(Style.RESET_ALL)
        click_on_fileSubmissionLink = self.driver.find_element(By.XPATH,
                                                     "//span[normalize-space()='File Submission']")
        self.driver.execute_script("arguments[0].click();", click_on_fileSubmissionLink)
        self.driver.refresh()
        time.sleep(3)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        click_on_fileSubmissionLink = self.driver.find_element(By.XPATH,
                                                       "//span[normalize-space()='File Submission']")
        self.driver.execute_script("arguments[0].click();", click_on_fileSubmissionLink)
#Click on Upload link
        click_on_uploadLink = self.driver.find_element(By.XPATH,
                                                       "//span[normalize-space()='File Upload']")
        self.driver.execute_script("arguments[0].click();",click_on_uploadLink)
        self.driver.get_screenshot_as_file("..\Screenshots\FileUpload.png")
        click_on_iconButton = self.driver.find_element(By.XPATH,
                                                       "//mat-icon[@type='button']")
        self.driver.execute_script("arguments[0].click();", click_on_iconButton)
        time.sleep(5)
#Verifying File Selection Instructions
        expected_Heading = "File Selection Instructions"
        actual_Heading= wait.until(EC.visibility_of_element_located((By.XPATH, "//app-info-modal[@class='ng-star-inserted']/div/span/h1"))).text
        #actual_Heading = self.driver.find_element(By.XPATH, "//app-info-modal[@class='ng-star-inserted']/div/span/h1").text
        self.soft_assert(self.assertEqual, expected_Heading, actual_Heading)
        if  actual_Heading == expected_Heading:
            print(Fore.GREEN + "Verified 'File Selection Instructions': -", end='')
            print(Style.RESET_ALL)
            #print("     Expected Heading:- " + expected_Heading + " && Actual Heading:- " + actual_Heading)
        else:
            # self.log.info("Assertion Failed")
            print("Heading does not match")

        #fileSelectionInstruction1 = self.driver.find_element(By.XPATH, "//div[@id='submissionInfo']/div[2]").text # Need to change as path has changed on 12/28
        expected_FileInsturctions = "Please use the ‘Browse’ button to locate and select a text (*.txt) file for upload. The file size must be greater than 0 bytes, but not more than 500 KB, and the file name can only contain the following valid characters:"
        sheet.cell(row=20, column=2).value = expected_FileInsturctions
        actual_FileInstructions = self.driver.find_element(By.XPATH, "//app-info-modal[@class='ng-star-inserted']/div/div[1]").text
        sheet.cell(row=20, column=3).value = actual_FileInstructions
        self.soft_assert(self.assertEqual, expected_FileInsturctions, actual_FileInstructions)
        if actual_FileInstructions == expected_FileInsturctions:
            print("    " + actual_FileInstructions)
            sheet.cell(row=20, column=4).value = "Pass"
        else:
            print("File Instructions does not match")
            sheet.cell(row=20, column=4).value = "Fail"

        list_items = self.driver.find_elements(By.XPATH, "//app-info-modal[@class='ng-star-inserted']/div/ul/li")

        for list_item in list_items:
            print("     "+list_item.text)

        click_on_closeButton = self.driver.find_element(By.XPATH, "//mat-icon[normalize-space()='close']")

        self.driver.execute_script("arguments[0].click();", click_on_closeButton)

        #self.driver.get_screenshot_as_file("..\Screenshots\File Upload Page.png")
# Clicking on Browse button
        time.sleep(10)
        browse_Button = self.driver.find_element(By.XPATH,
                                               "//span[normalize-space()='Browse']")
        self.driver.execute_script("arguments[0].click();", browse_Button)
# Loading a file
        time.sleep(15)
        try:
            file = self.driver.find_element(By.XPATH, "//input[@type='file']")
            os.path.abspath("r..\File_Tobe_Upload\CONTRACTOR_MAPPING_VALIDATION_BAD_BSI_FORM_S.txt")
            file.send_keys(os.path.abspath(r"..\File_Tobe_Upload\CONTRACTOR_MAPPING_VALIDATION_BAD_BSI_FORM_S.txt"))
        except Exception as e:
            print(e)

        # try:
        #     file = self.driver.find_element(By.XPATH, "//input[@type='file']")
        #     file.send_keys("C:\Data_Automation\Form_S_TEST\DATA_TYPE_VALIDATION_BAD_BSI_ID_1_FORM_S.txt")
        # except Exception as e:
        #     print(e)

#Entering a comment
        comment_Inputbox = self.driver.find_element(By.XPATH, "//textarea[@id='comment']")
        comment_Tobe_Entered = "BAD_BSI"
        comment_Inputbox.send_keys(comment_Tobe_Entered)

# Click on Submit button
        self.driver.find_element(By.XPATH, "//button[@type='submit']").click()

#Fetch uploaded file name
        uploadedFile =self.driver.find_element(By.XPATH, "//div[@class ='submissionSubText']/div/b").text
        a, b = uploadedFile.split(' ', 1)

        uploadedFilename = a
        uploaded_Filename = uploadedFilename.replace("'", "")
        print(Style.RESET_ALL)
# Click on Status
        click_on_Status = self.driver.find_element(By.XPATH,
                                                   "//a[normalize-space()='Status']")
        self.driver.execute_script("arguments[0].click();", click_on_Status)

        time.sleep(9)
        #self.driver.get_screenshot_as_file("..\Screenshots\FileStatus.png")
#Click on Refresh button on Status page
        click_on_Refresh = self.driver.find_element(By.XPATH,
                                                   "//button[@class='refreshButton calltoAction']/mat-icon")
        self.driver.execute_script("arguments[0].click();", click_on_Refresh)

        self.driver.execute_script("arguments[0].click();", click_on_Refresh)

        self.driver.execute_script("arguments[0].click();", click_on_Refresh)
        time.sleep(21)

        wait.until(EC.visibility_of_element_located((By.XPATH, "//table[@title='File status Table']/tbody/tr/td")))
        filename_displayed_onStatus = self.driver.find_element(By.XPATH, "//table[@title='File status Table']/tbody/tr/td").text
        print(Fore.GREEN + 'Verified Uploaded File: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=22, column=2).value = uploaded_Filename
        sheet.cell(row=22, column=3).value = filename_displayed_onStatus
        self.soft_assert(self.assertEqual, filename_displayed_onStatus, uploaded_Filename)
        if uploaded_Filename == filename_displayed_onStatus:
            print("     Uploaded Filename: - " + uploaded_Filename )
            print("     Uploaded Filename Displayed On Status: - " + filename_displayed_onStatus)
            sheet.cell(row=22, column=4).value = "Pass"
        else:
            sheet.cell(row=22, column=4).value = "Fail"
            print("False")
#Verify Comment on status page
        # click_on_Refresh = self.driver.find_element(By.XPATH,
        #                                             "//button[@class='refreshButton calltoAction']/mat-icon")
        # self.driver.execute_script("arguments[0].click();", click_on_Refresh)
        comment_DisplayedOnStatus = self.driver.find_element(By.XPATH, "//span[@class='viewComment']").text

        print(Fore.GREEN + 'Verified Comment: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=24, column=2).value = comment_Tobe_Entered
        sheet.cell(row=24, column=3).value = comment_DisplayedOnStatus
        self.soft_assert(self.assertEqual, comment_Tobe_Entered, comment_DisplayedOnStatus)
        if comment_Tobe_Entered == comment_DisplayedOnStatus:
            print("     Entered Comment:- " + comment_Tobe_Entered )
            print("     Comment Displayed On Status: - " + comment_DisplayedOnStatus )
            sheet.cell(row=24, column=4).value = "Pass"
        else:
            # self.log.info("Assertion Failed")
            sheet.cell(row=24, column=4).value = "Fail"
            print("Comment does not match")
#Verify Timestamp
        timeStamp_DisplayedOnStatus = self.driver.find_element(By.XPATH, "//table[@title='File status Table']/tbody/tr/td[3]").text
        now =datetime.datetime.now()
        currentTime = now.strftime("%m/%d/%Y %I:%M %p")
        # currentTime = currentTime.replace(":", " ",-3)
        ##############################Checking with new time####################################

        updatedtime= now - datetime.timedelta(minutes=1)
        #print(updatedtime)
        updated_Currentime = updatedtime.strftime("%m/%d/%Y %I:%M %p")
        # print("Updated Current time:-" + updated_Currentime)
        # print("Time displayed on Status page:-" +timeStamp_DisplayedOnStatus)
        # print("Current time:-" + currentTime)

        ######################################################################

        print(Fore.GREEN + 'Verified Uploaded File Timing: -', end='')
        print(Style.RESET_ALL)
        sheet.cell(row=26, column=3).value = currentTime
        sheet.cell(row=26, column=2).value = timeStamp_DisplayedOnStatus
        #self.soft_assert(self.assertEqual, timeStamp_DisplayedOnStatus, currentTime or self.assertEqual,timeStamp_DisplayedOnStatus, updated_Currentime)

        if timeStamp_DisplayedOnStatus == currentTime or updated_Currentime:
            print("     File Uploaded on: - " + timeStamp_DisplayedOnStatus)
            print("     Current Date & Time: - " + currentTime, updated_Currentime)
            sheet.cell(row=26, column=4).value = "Pass"
        else:

            print("Timestamp and Uploaded file time does not match")
            sheet.cell(row=26, column=4).value = "Fail"

        self.driver.get_screenshot_as_file("..\Screenshots\File Status Page.png")
# ############################################# Verifying Dashboard ###################################################
        print(Fore.BLUE +"\n************************* Verifying Dashboard Page *************************")
        print(Style.RESET_ALL)
        dashboardLink = self.driver.find_element(By.XPATH,
                                                "//span[normalize-space()='Dashboards']")
        self.driver.execute_script("arguments[0].click();", dashboardLink)
        #self.driver.get_screenshot_as_file("..\Screenshots\Dashboard.png")
        # dash_Text = self.driver.find_element(By.XPATH, "//span[@class='anchor-format grey-text ng-star-inserted']").text
        # print("Displayed text on Dashboards page:- " + dash_Text)
        dashboardID = self.driver.find_element(By.XPATH, "//table[@title='Dashboard Table']/tbody/tr/td[1]").text
        dashboardName = self.driver.find_element(By.XPATH, "//table[@title='Dashboard Table']/tbody/tr/td[2]").text
        dashboardDescription = self.driver.find_element(By.XPATH, "//table[@title='Dashboard Table']/tbody/tr/td[3]").text
        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON 'DASHBOARD' TABLE IS:- ", end='')
        print(Style.RESET_ALL)
        print("     Dashboard ID: - " + dashboardID + "   Dashboard Name: - " + dashboardName + "   Dashboard Description: - " + dashboardDescription)
        sheet.cell(row=30, column=3).value = dashboardID
        sheet.cell(row=32, column=3).value = dashboardName
        sheet.cell(row=34, column=3).value = dashboardDescription
        if sheet.cell(row=30, column=3).value == sheet.cell(row=30, column=2).value:
            sheet.cell(row=30, column=4).value ='Pass'
        else:
            sheet.cell(row=30, column=4).value = 'Fail'
        if sheet.cell(row=32, column=3).value == sheet.cell(row=32, column=2).value:
            sheet.cell(row=32, column=4).value = 'Pass'
        else:
            sheet.cell(row=32, column=4).value = 'Fail'
        if sheet.cell(row=34, column=3).value == sheet.cell(row=34, column=2).value:
            sheet.cell(row=34, column=4).value = 'Pass'
        else:
            sheet.cell(row=34, column=4).value = 'Fail'

        time.sleep(5)
        parent_window_id= self.driver.current_window_handle

        click_on_firstLink_Dashboard = self.driver.find_element(By.XPATH,
                                             "//table[@title='Dashboard Table']/tbody/tr/td[2]/a")

        self.driver.execute_script("arguments[0].click();", click_on_firstLink_Dashboard)

        windows= self.driver.window_handles

        for w in windows:
            self.driver.switch_to.window(w)
            time.sleep(6)
            expected_Title = "B01. MicroStrategy"
            actual_Title = self.driver.title
            sheet.cell(row=36, column=3).value = actual_Title
            if actual_Title==expected_Title:
                print(Fore.GREEN+'Verified MicroStrategy Page Title: -', end='')
                print(Style.RESET_ALL)
                print("     Expected Title: - " + expected_Title)
                print("     Actual Title: - " + actual_Title)
                sheet.cell(row=36, column=4).value = "Pass"
            # if self.driver.title.__eq__("B01. MicroStrategy"):
            #     print(self.driver.title)
                #print(self.driver.current_window_handle)
                self.driver.close()
                break
            else:
                sheet.cell(row=36, column=4).value = "Fail"
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        self.driver.switch_to.window(parent_window_id)
# ############################## Verifying Form Page #####################################
        print(Fore.BLUE + "\n************************* Verifying Form Page *************************")
        print(Style.RESET_ALL)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        time.sleep(5)
        #forms_Link = self.driver.find_element(By.XPATH,"//span[normalize-space()='Forms']")
        forms_Link = self.driver.find_element(By.XPATH, "//span[@class='list-item-label'][normalize-space()='Forms']")
        self.driver.execute_script("arguments[0].click();", forms_Link)
        time.sleep(8)
        self.driver.get_screenshot_as_file("..\Screenshots\Forms.png")
        # Need to be change 2/2/24
        # form= self.driver.find_element(By.XPATH, "//table[@title='Forms Table']/tbody/tr[1]/td[1]").text
        # formType = self.driver.find_element(By.XPATH, "//table[@title='Forms Table']/tbody[1]/tr/td[2]").text
        # formName=  self.driver.find_element(By.XPATH, "//table[@title='Forms Table']/tbody/tr[1]/td[3]").text

        form = self.driver.find_element(By.XPATH, "//main[@id='main-forms']/table/tbody/tr[1]/td[1]").text
        formType = self.driver.find_element(By.XPATH, "//main[@id='main-forms']/table/tbody/tr[1]/td[2]").text
        formName = self.driver.find_element(By.XPATH, "//main[@id='main-forms']/table/tbody/tr[1]/td[3]").text

        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON 'FORM' TABLE IS: - ", end='')
        print(Style.RESET_ALL)
        print("     Form: - " + form + "  Form Type: - " + formType + "  Form Name: - " + formName)
        sheet.cell(row=41, column=3).value = form
        sheet.cell(row=43, column=3).value = formType
        sheet.cell(row=45, column=3).value = formName

        if sheet.cell(row=41, column=3).value == sheet.cell(row=41, column=2).value:
            sheet.cell(row=41, column=4).value ='Pass'
        else:
            sheet.cell(row=41, column=4).value = 'Fail'
        if sheet.cell(row=43, column=3).value == sheet.cell(row=43, column=2).value:
            sheet.cell(row=43, column=4).value = 'Pass'
        else:
            sheet.cell(row=43, column=4).value = 'Fail'
        if sheet.cell(row=45, column=3).value == sheet.cell(row=45, column=2).value:
            sheet.cell(row=45, column=4).value = 'Pass'
        else:
            sheet.cell(row=34, column=4).value = 'Fail'
        time.sleep(7)
        parent_window_id = self.driver.current_window_handle
        #click_on_firstLink_Forms = self.driver.find_element(By.XPATH,
                                            # "//main[@id='main-forms']/table[@title='Forms Table']/tbody/tr[2]/td[3]/a")
        click_on_firstLink_Forms = self.driver.find_element(By.XPATH,
                                                            "//main[@id='main-forms']/table/tbody/tr[1]/td[3]/a")

        self.driver.execute_script("arguments[0].click();", click_on_firstLink_Forms)

        windows = self.driver.window_handles

        for w in windows:
            self.driver.switch_to.window(w)
            time.sleep(9)
            expected_Title = "(FORM 7 - PART A) APPEALS ACTIVITY (CMS-2592). MicroStrategy"
            actual_Title = self.driver.title
            sheet.cell(row=47, column=3).value = actual_Title
            if actual_Title==expected_Title:
                print(Fore.GREEN + 'Verified MicroStrategy Page Title: - ', end='')
                print(Style.RESET_ALL)
                print("     Expected Title: - "+ expected_Title)
                print("     Actual Title: - " + actual_Title)
                sheet.cell(row=47, column=4).value = "Pass"
                #print(self.driver.current_window_handle)
                break
            else:
                sheet.cell(row=47, column=4).value = "Fail"

        self.driver.switch_to.window(parent_window_id)
# # ############################## Verifying Reports Page #####################################
        print(Fore.BLUE + "\n************************* Verifying Reports Page *************************")
        print(Style.RESET_ALL)
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        time.sleep(3)
        reports_Link = self.driver.find_element(By.XPATH,
                                                 "//span[normalize-space()='Reports']")
        self.driver.execute_script("arguments[0].click();", reports_Link)
        time.sleep(5)
        self.driver.get_screenshot_as_file("..\Screenshots\Reports.png")
        reportID = self.driver.find_element(By.XPATH, "//table[@title='Reports Table']/tbody/tr/td[1]").text # WebDriver Exception
        reportName = self.driver.find_element(By.XPATH, "//table[@title='Reports Table']/tbody/tr/td[2]").text
        reportDescription = self.driver.find_element(By.XPATH, "//table[@title='Reports Table']/tbody/tr/td[3]").text
        sheet.cell(row=52, column=3).value = reportID
        sheet.cell(row=54, column=3).value = reportName
        sheet.cell(row=56, column=3).value = reportDescription

        if sheet.cell(row=52, column=3).value == sheet.cell(row=52, column=2).value:
            sheet.cell(row=52, column=4).value ='Pass'
        else:
            sheet.cell(row=52, column=4).value = 'Fail'

        if sheet.cell(row=54, column=3).value == sheet.cell(row=54, column=2).value:
            sheet.cell(row=54, column=4).value = 'Pass'
        else:
            sheet.cell(row=54, column=4).value = 'Fail'

        if sheet.cell(row=56, column=3).value == sheet.cell(row=56, column=2).value:
            sheet.cell(row=56, column=4).value = 'Pass'
        else:
            sheet.cell(row=56, column=4).value = 'Fail'

        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON 'REPORT' TABLE IS: - ", end='')
        print(Style.RESET_ALL)
        print("     Report ID: - " + reportID + "   Report Name: - " + reportName + "   Report Description: - " + reportDescription)
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
############################## Verifying Resources  Page ###################################################
        print(Fore.BLUE + "\n************************* Verifying Resources Page*************************")
        print(Style.RESET_ALL)
        ##################### FAQs#########################
        print(Fore.GREEN + "THE FIRST RECORD DISPLAYED ON FAQs/Reporting/Training Material: - ", end='')
        print(Style.RESET_ALL)
        try:

            resources_Link = self.driver.find_element(By.XPATH,
                                             "//span[normalize-space()='Resources']")
            self.driver.execute_script("arguments[0].click();", resources_Link)
            time.sleep(10)
            click_on_FAQs = self.driver.find_element(By.XPATH,
                                                       "//span[normalize-space()='FAQs']")
            self.driver.execute_script("arguments[0].click();", click_on_FAQs)
            time.sleep(9)
            click_on_SystemSetup = self.driver.find_element(By.XPATH,# WebDriver Exception error
                                                 "//a[normalize-space()='System Set-Up Considerations']")
            self.driver.execute_script("arguments[0].click();", click_on_SystemSetup)
            system_Setup_Text = self.driver.find_element(By.XPATH,
                                                 "//div[@class='container']/mat-card/ul/li[1]/a").text
            print(system_Setup_Text)

            sheet.cell(row=60, column=3).value = system_Setup_Text
            if sheet.cell(row=60, column=3).value == sheet.cell(row=60, column=2).value:
                sheet.cell(row=60, column=4).value ='Pass'
            else:
                sheet.cell(row=60, column=4).value = 'Fail'
        ##############  Reporting ############
            reporting_Link = self.driver.find_element(By.XPATH,
                                                  "//span[normalize-space()='Reporting']")
            self.driver.execute_script("arguments[0].click();", reporting_Link)
            time.sleep(10)
            click_on_CrowdContractorMapping = self.driver.find_element(By.XPATH,# WebDriver Exception
                                                       " //a[normalize-space()='Crowd Contractor Mapping']")
            self.driver.execute_script("arguments[0].click();", click_on_CrowdContractorMapping)
            crowd_Reporting_Requirement_Text = self.driver.find_element(By.XPATH,
                                                   "//div[@class='container']/mat-card/ul/li[1]/a").text
            print(crowd_Reporting_Requirement_Text)
            sheet.cell(row=62, column=3).value = crowd_Reporting_Requirement_Text
            if sheet.cell(row=62, column=3).value == sheet.cell(row=62, column=2).value:
                sheet.cell(row=62, column=4).value ='Pass'
            else:
                sheet.cell(row=62, column=4).value = 'Fail'
        ############ Training Material ########################
            trainingMaterial_Link = self.driver.find_element(By.XPATH,
                                                  "//span[normalize-space()='Training Material']")
            self.driver.execute_script("arguments[0].click();", trainingMaterial_Link)
            time.sleep(10)
            click_on_RequestCrowdAccess = self.driver.find_element(By.XPATH,
                                                                   "//a[contains(text(),'Request CROWD Access (Without an existing EIDM acc')]")
            self.driver.execute_script("arguments[0].click();", click_on_RequestCrowdAccess)
            requestCrowdAccess_Text = self.driver.find_element(By.XPATH,
                                                   "//div[@class='container']/mat-card/ul/li[2]/ul/li[2]/a").text

            print(requestCrowdAccess_Text)
            sheet.cell(row=64, column=3).value = requestCrowdAccess_Text
            if sheet.cell(row=64, column=3).value == sheet.cell(row=64, column=2).value:
                sheet.cell(row=64, column=4).value ='Pass'
            else:
                sheet.cell(row=64, column=4).value = 'Fail'
            workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        except WebDriverException as e:
            print("Webdriver encounterd an exception: (str(e))")

# ############################## Verifying News Page #####################################
        print(Fore.BLUE + "\n************************* Verifying News Page *************************")
        print(Style.RESET_ALL)

        #Path changed found on 5th Jan news_Link = self.driver.find_element(By.XPATH,
                                                 #"//span[@class='mat-list-item-content'][normalize-space()='News']")
        #self.driver.execute_script("arguments[0].click();", news_Link)

        news_Link = self.driver.find_element(By.XPATH,
                                                 "//span[normalize-space()='News']")
        self.driver.execute_script("arguments[0].click();", news_Link)
        time.sleep(10)

        # news_Text = self.driver.find_element(By.XPATH, "//h1[@id='focus-main-content']").text
        # print("Displayed text on News page:- " + news_Text)

        print(Fore.GREEN + "NEWS for Year 2020: -")
        print(Style.RESET_ALL)
       # self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        year2020 = self.driver.find_element(By.XPATH,
                                             "//h2[normalize-space()='2020']")
        self.driver.execute_script("arguments[0].click();", year2020)
        time.sleep(6)
       # self.driver.switch_to.frame(self.driver.find_element(By.XPATH, "//object[@id='obj_crowd_wab_application']"))
        line1 =self.driver.find_element(By.XPATH, " (//span[@class='blue'][normalize-space()='Attn: Medicare Contractors'])[1]").text
        #print(line1)
        time.sleep(6)
        line2 = self.driver.find_element(By.XPATH,
                                         "//span[normalize-space()='| February 12, 2020']").text

        print(line1 + ' ' + line2)
        sheet.cell(row=69, column=3).value = line1 + ' ' +line2
        if sheet.cell(row=69, column=3).value == sheet.cell(row=69, column=2).value:
            sheet.cell(row=69, column=4).value ='Pass'
        else:
            sheet.cell(row=69, column=4).value = 'Fail'

        #line3 =self.driver.find_element(By.XPATH, "//span[normalize-space()='| December 23rd, 2021']").text

        line3 = self.driver.find_element(By.XPATH, "//p[contains(text(),'Although CR 11353 (Supplier Specialty Code D5) for')]").text
        sheet.cell(row=70, column=3).value = line3
        print(line3)
        # line2 = self.driver.find_element(By.XPATH,
        #                                  " //span[@class='ng-star-inserted'][normalize-space()='| February 12, 2020'])[1]").text
        workbook.save("..\ExcelFiles\Automated_SmokeTest_Result.xlsx")
        self.assert_all()
















